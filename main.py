import os
import sys
import asyncio
import logging
from io import BytesIO
from statistics import mean
from time import time
from datetime import datetime
import contextlib

import pytz
import aiohttp
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.types import (
    Message, ReplyKeyboardMarkup, KeyboardButton, Update
)

from aiohttp import web
from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ---------------- ENV & CONFIG ----------------
load_dotenv()
TOKEN = os.getenv("TELEGRAM_TOKEN")
BASE_URL = os.getenv("BASE_URL", "").rstrip("/")
TZ = os.getenv("TZ", "Europe/Stockholm")
VERSION = "v0.6-webhook"

if not TOKEN:
    raise RuntimeError("TELEGRAM_TOKEN is not set")
if not BASE_URL.startswith("https://"):
    raise RuntimeError("BASE_URL must be your public https Render URL, e.g. https://<service>.onrender.com")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

bot = Bot(TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

# ---------------- CONSTANTS ----------------
BYBIT_API = "https://api.bybit.com"
REQUEST_TIMEOUT = 25
HTTP_HEADERS = {"User-Agent": "InnertradeScreener/1.0 (+render.com)"}

# Base universe (can be extended by /add SYMBOL)
SYMBOLS_BYBIT = [
    "BTCUSDT","ETHUSDT","SOLUSDT","XRPUSDT","BNBUSDT",
    "DOGEUSDT","ADAUSDT","LINKUSDT","TRXUSDT","TONUSDT","ARBUSDT","OPUSDT"
]

# ---------------- RUNTIME STATE ----------------
USERS: dict[int, dict] = {}
DEFAULT_USER = {
    "exchange": "bybit",
    "mode": "active",      # "active" | "passive"
    "quiet": False,         # night mode placeholder
    "watchlist": [],
    "alerts": {
        "vol_mult": 1.5,   # 5m turnover vs MA20
        "pct24_abs": 3.0,  # |24h %|
        "cooldown_min": 20
    },
    "last_alert": {}        # symbol -> ts
}

def ensure_user(uid: int) -> dict:
    if uid not in USERS:
        st = {k: (v.copy() if isinstance(v, dict) else (v[:] if isinstance(v, list) else v))
              for k, v in DEFAULT_USER.items()}
        st["watchlist"] = []
        st["last_alert"] = {}
        USERS[uid] = st
    return USERS[uid]

# ---------------- SIMPLE CACHE ----------------
CACHE: dict[str, tuple[float, str]] = {}

def cache_get(key: str, ttl: int):
    item = CACHE.get(key)
    if not item:
        return None
    ts, val = item
    return val if (time() - ts) < ttl else None

def cache_set(key: str, val: str):
    CACHE[key] = (time(), val)

# ---------------- MARKET HEADER (stub) ----------------
async def render_header_text() -> str:
    btc_dom = "54.1% (+0.3)"
    funding = "+0.012%"
    fg = "34 (-3)"
    return f"🧭 <b>Market mood</b>\nBTC.D: {btc_dom} | Funding avg: {funding} | F&G: {fg}"

# ---------------- HTTP HELPERS ----------------
async def http_get_json(session: aiohttp.ClientSession, url: str, params: dict | None = None):
    for attempt in range(3):
        try:
            async with session.get(url, params=params, headers=HTTP_HEADERS, timeout=REQUEST_TIMEOUT) as r:
                text = await r.text()
                if r.status != 200:
                    logging.warning(f"[HTTP {r.status}] {url} {params} -> {text[:180]}")
                    await asyncio.sleep(0.6)
                    continue
                try:
                    return await r.json()
                except Exception as e:
                    logging.warning(f"[JSON ERR] {url} -> {e}")
                    await asyncio.sleep(0.6)
        except Exception as e:
            logging.warning(f"[REQ ERR] {url} {params} -> {e}")
            await asyncio.sleep(0.8)
    return None

Iurii Qechunts, [05.09.2025 23:54]
# ---------------- BYBIT PROVIDERS ----------------
async def bybit_klines(session: aiohttp.ClientSession, symbol: str, interval_minutes: int, limit: int):
    interval = str(interval_minutes) if interval_minutes in (1,3,5,15,30,60,120,240,360,720) else "5"
    url = f"{BYBIT_API}/v5/market/kline"
    params = {"category": "linear", "symbol": symbol, "interval": interval, "limit": str(limit)}
    return await http_get_json(session, url, params)

async def bybit_ticker(session: aiohttp.ClientSession, symbol: str):
    url = f"{BYBIT_API}/v5/market/tickers"
    params = {"category": "linear", "symbol": symbol}
    return await http_get_json(session, url, params)

def parse_bybit_row(row):
    # [start, open, high, low, close, volume, turnover]
    o = float(row[1]); h = float(row[2]); l = float(row[3]); c = float(row[4])
    v = float(row[5]); turnover = float(row[6])
    return o, h, l, c, v, turnover

# ---------------- INDICATORS ----------------

def moving_average(values, length):
    if not values or len(values) < length:
        return None
    return mean(values[-length:])


def compute_atr(ohlc_rows, period: int = 14):
    if len(ohlc_rows) < period + 1:
        return None
    trs = []
    prev_close = ohlc_rows[0][3]
    for i in range(1, len(ohlc_rows)):
        _, h, l, c, *_ = ohlc_rows[i]
        tr = max(h - l, abs(h - prev_close), abs(prev_close - l))
        trs.append(tr)
        prev_close = c
    if len(trs) < period:
        return None
    return mean(trs[-period:])


def slope(values, lookback: int = 10):
    if len(values) < 2 * lookback:
        return 0.0
    return mean(values[-lookback:]) - mean(values[-2*lookback:-lookback])

# ---------------- SCREENER BUILDERS ----------------

async def build_activity_bybit(session: aiohttp.ClientSession) -> list[dict]:
    # Activity = turnover spike (5m vs MA20) + 24h share in 7d
    async def fetch(sym: str):
        try:
            k5 = await bybit_klines(session, sym, 5, 220)
            k1h = await bybit_klines(session, sym, 60, 180)
            return sym, k5, k1h
        except Exception as e:
            logging.warning(f"[fetch_act ERR] {sym} -> {e}")
            return sym, None, None

    raw = await asyncio.gather(*[asyncio.create_task(fetch(s)) for s in SYMBOLS_BYBIT])
    out = []
    for sym, k5, k1h in raw:
        try:
            l5 = (k5 or {}).get("result", {}).get("list", [])
            l1h = (k1h or {}).get("result", {}).get("list", [])
            if len(l5) < 40 or len(l1h) < 50:
                continue
            rows5 = [parse_bybit_row(r) for r in l5]
            rows1h = [parse_bybit_row(r) for r in l1h]
            turns5 = [r[5] for r in rows5]
            ma20 = moving_average(turns5[:-1], 20)
            if not ma20 or ma20 <= 0:
                continue
            vol_mult = turns5[-1] / ma20

            turns1h = [r[5] for r in rows1h]
            vol_24h = sum(turns1h[-24:])
            vol_7d  = sum(turns1h[-168:]) if len(turns1h) >= 168 else max(sum(turns1h), 1.0)
            share24 = int(round((vol_24h / vol_7d) * 100)) if vol_7d > 0 else 0

            out.append({"symbol": sym, "venue": "Bybit", "vol_mult": vol_mult, "share24": share24})
        except Exception as e:
            logging.warning(f"[build_activity parse ERR] {sym} -> {e}")
    out.sort(key=lambda x: x["vol_mult"], reverse=True)
    return out[:10]


async def build_volatility_bybit(session: aiohttp.ClientSession) -> list[dict]:
    # Volatility = ATR% (5m) + current turnover spike vs MA20
    async def fetch(sym: str):
        try:
            k5 = await bybit_klines(session, sym, 5, 400)
            return sym, k5
        except Exception as e:
            logging.warning(f"[fetch_vol ERR] {sym} -> {e}")
            return sym, None

    raws = await asyncio.gather(*[asyncio.create_task(fetch(s)) for s in SYMBOLS_BYBIT])

Iurii Qechunts, [05.09.2025 23:54]
out = []
    for sym, k5 in raws:
        try:
            l5 = (k5 or {}).get("result", {}).get("list", [])
            if len(l5) < 100:
                continue
            rows5 = [parse_bybit_row(r) for r in l5]
            closes = [r[3] for r in rows5]
            turns  = [r[5] for r in rows5]
            last_close = closes[-1]
            atr_val = compute_atr(rows5, 14)
            if not atr_val or last_close <= 0:
                continue
            atr_pct = atr_val / last_close * 100.0
            ma20 = moving_average(turns[:-1], 20)
            vol_mult = (turns[-1] / ma20) if ma20 else 0.0
            out.append({"symbol": sym, "venue": "Bybit", "atr_pct": atr_pct, "vol_mult": vol_mult})
        except Exception as e:
            logging.warning(f"[build_vol parse ERR] {sym} -> {e}")
    out.sort(key=lambda x: x["atr_pct"], reverse=True)
    return out[:10]


async def build_trend_bybit(session: aiohttp.ClientSession) -> list[dict]:
    # Trend = relation to MA200/MA360 + slope
    async def fetch(sym: str):
        try:
            k5 = await bybit_klines(session, sym, 5, 400)
            return sym, k5
        except Exception as e:
            logging.warning(f"[fetch_trend ERR] {sym} -> {e}")
            return sym, None

    raws = await asyncio.gather(*[asyncio.create_task(fetch(s)) for s in SYMBOLS_BYBIT])
    res = []
    for sym, k5 in raws:
        try:
            l5 = (k5 or {}).get("result", {}).get("list", [])
            if len(l5) < 380:
                continue
            rows5 = [parse_bybit_row(r) for r in l5]
            closes = [r[3] for r in rows5]
            ma200 = moving_average(closes, 200)
            ma360 = moving_average(closes, 360)
            if ma200 is None or ma360 is None:
                continue
            last_close = closes[-1]
            above200 = last_close > ma200
            above360 = last_close > ma360
            s200 = slope(closes[-220:], 10)
            # ATR momentum as volatility change proxy
            atr_now = compute_atr(rows5, 14)
            atr_prev = compute_atr(rows5[-60:-15], 14) if len(rows5) > 75 else None
            vol_change = "↑" if (atr_prev and atr_now and atr_now > atr_prev) else ("↓" if atr_prev else "≈")
            res.append({"symbol": sym, "venue":"Bybit", "above200": above200, "above360": above360, "slope200": s200, "vol_change": vol_change})
        except Exception as e:
            logging.warning(f"[build_trend parse ERR] {sym} -> {e}")
    res.sort(key=lambda x: (x["above200"], x["above360"], x["slope200"]), reverse=True)
    return res[:10]

# ---------------- RENDERERS ----------------
async def render_activity_text() -> str:
    key = "activity:bybit"
    cached = cache_get(key, ttl=45)
    if cached:
        return cached
    async with aiohttp.ClientSession() as s:
        items = await build_activity_bybit(s)
    if not items:
        txt = "\n🔥 <b>Активность</b>\nНет данных (тихо/таймаут/лимиты)."
        cache_set(key, txt)
        return txt
    lines = ["\n🔥 <b>Активность</b> (Bybit)"]
    for i, r in enumerate(items, 1):
        lines.append(f"{i}) {r['symbol']} ({r['venue']}) Vol x{r['vol_mult']:.1f} | 24h vs 7d: {r['share24']}%")
    txt = "\n".join(lines)
    cache_set(key, txt)
    return txt


async def render_volatility_text() -> str:
    key = "vol:bybit"
    cached = cache_get(key, ttl=60)
    if cached:
        return cached
    async with aiohttp.ClientSession() as s:
        items = await build_volatility_bybit(s)
    if not items:
        txt = "\n⚡ <b>Волатильность</b>\nНет данных."
        cache_set(key, txt)
        return txt
    lines = ["\n⚡ <b>Волатильность</b> (ATR%, 5m, Bybit)"]
    for i, r in enumerate(items, 1):
        lines.append(f"{i}) {r['symbol']} ({r['venue']}) ATR {r['atr_pct']:.2f}% | Vol x{r['vol_mult']:.1f}")
    txt = "\n".join(lines)
    cache_set(key, txt)
    return txt

Iurii Qechunts, [05.09.2025 23:54]
async def render_trend_text() -> str:
    key = "trend:bybit"
    cached = cache_get(key, ttl=60)
    if cached:
        return cached
    async with aiohttp.ClientSession() as s:
        items = await build_trend_bybit(s)
    if not items:
        txt = "\n📈 <b>Тренд</b>\nНет данных."
        cache_set(key, txt)
        return txt
    lines = ["\n📈 <b>Тренд</b> (5m, MA200/MA360, Bybit)"]
    for i, r in enumerate(items, 1):
        pos = [">MA200" if r["above200"] else "<MA200", ">MA360" if r["above360"] else "<MA360"]
        slope_tag = "slope+" if r["slope200"] > 0 else ("slope-" if r["slope200"] < 0 else "slope≈")
        lines.append(f"{i}) {r['symbol']} ({r['venue']}) {' & '.join(pos)} | {slope_tag} | вола {r['vol_change']}")
    txt = "\n".join(lines)
    cache_set(key, txt)
    return txt

# ---------------- BUBBLES ----------------
async def build_bubbles_data():
    async with aiohttp.ClientSession() as s:
        raws = await asyncio.gather(*[asyncio.create_task(bybit_ticker(s, sym)) for sym in SYMBOLS_BYBIT])
    out = []
    for sym, data in zip(SYMBOLS_BYBIT, raws):
        try:
            lst = (data or {}).get("result", {}).get("list", [])
            if not lst:
                continue
            it = lst[0]
            pct = float(it.get("price24hPcnt", 0.0)) * 100.0
            turnover = float(it.get("turnover24h", 0.0))
            out.append({"symbol": sym, "pct": pct, "turnover": turnover})
        except Exception:
            continue
    return out


def render_bubbles_png(items) -> bytes:
    buf = BytesIO()
    if not items:
        fig = plt.figure(figsize=(8,4), dpi=160)
        ax = fig.add_subplot(111); ax.axis("off")
        ax.text(0.5,0.5,"Нет данных для пузырьков", ha="center", va="center", fontsize=16)
        fig.savefig(buf, format="png"); plt.close(fig)
        return buf.getvalue()

    turnovers = np.array([max(1.0, it["turnover"]) for it in items], dtype=float)
    sizes = np.sqrt(turnovers)
    k = (8000.0 / sizes.max()) if sizes.max() > 0 else 1.0
    s = sizes * k

    n = len(items)
    cols = int(np.ceil(np.sqrt(n))); rows = int(np.ceil(n / cols))
    xs, ys = [], []
    for i in range(n):
        r = i // cols; c = i % cols
        xs.append(c); ys.append(rows - 1 - r)
    xs = np.array(xs); ys = np.array(ys)

    colors = []
    for it in items:
        if it["pct"] > 0.5: colors.append("#16a34a")
        elif it["pct"] < -0.5: colors.append("#dc2626")
        else: colors.append("#6b7280")
    labels = [f"{it['symbol']}\n{it['pct']:+.1f}%" for it in items]

    fig = plt.figure(figsize=(12,7), dpi=160)
    ax = fig.add_subplot(111)
    ax.set_facecolor("#0b1020"); fig.patch.set_facecolor("#0b1020")
    ax.scatter(xs, ys, s=s, c=colors, alpha=0.85)
    for x, y, lab in zip(xs, ys, labels):
        ax.text(x, y, lab, ha="center", va="center", color="white", fontsize=9, weight="bold")
    ax.set_xticks([]); ax.set_yticks([])
    ax.set_xlim(-0.8, cols-0.2); ax.set_ylim(-0.2, rows-0.2)
    ax.set_title("Daily Bubbles (Bybit, 24h % | size ~ turnover)", color="white", fontsize=14)
    fig.tight_layout()
    fig.savefig(buf, format="png", facecolor=fig.get_facecolor(), bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


async def render_bubbles_message():
    data = await build_bubbles_data()
    header = await render_header_text()
    png = render_bubbles_png(data)
    return header, png

# ---------------- EXCEL CALCULATOR ----------------
async def build_risk_excel_template()->bytes:
    wb = Workbook(); ws = wb.active; ws.title = "RiskCalc"
    headers = ["Equity (USDT)", "Risk %", "Side", "Entry", "Stop",
               "Risk $", "Stop Dist", "Qty", "Leverage", "Notional $", "TP1", "TP2", "TP3"]
    ws.append(headers)
    ws["A2"]=10000; ws["B2"]=0.01; ws["C2"]="LONG"; ws["D2"]=100.0; ws["E2"]=95.0; ws["I2"]=5

Iurii Qechunts, [05.09.2025 23:54]
ws["F2"]="=A2*B2"
    ws["G2"]='=ABS(IF(UPPER(C2)="LONG",D2-E2,E2-D2))'
    ws["H2"]="=IF(G2>0,F2/G2,0)"
    ws["J2"]="=H2*D2"
    ws["K2"]='=IF(UPPER(C2)="LONG",D2+G2,D2-G2)'
    ws["L2"]='=IF(UPPER(C2)="LONG",D2+2*G2,D2-2*G2)'
    ws["M2"]='=IF(UPPER(C2)="LONG",D2+3*G2,D2-3*G2)'

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="1f2937")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="404040")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = white; cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for col in "ABCDEFGHIJKLM":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["A"].width = 16

    bio = BytesIO(); wb.save(bio); return bio.getvalue()

# ---------------- KEYBOARDS ----------------

def bottom_menu_kb()->ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Активность"), KeyboardButton(text="⚡ Волатильность")],
            [KeyboardButton(text="📈 Тренд"),      KeyboardButton(text="🫧 Bubbles")],
            [KeyboardButton(text="📰 Новости"),    KeyboardButton(text="🧮 Калькулятор")],
            [KeyboardButton(text="⭐ Watchlist"),   KeyboardButton(text="⚙️ Настройки")],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Выберите раздел…",
    )


def settings_text(u: dict) -> str:
    a = u.get("alerts", {})
    return (
        "Настройки:\n"
        f"Биржа: {u.get('exchange','bybit')}\n"
        f"Режим: {u.get('mode')}\n"
        f"Тихие часы: {u.get('quiet')}\n"
        f"Алерты: Vol x≥{a.get('vol_mult',1.5):.1f} | |24h%|≥{a.get('pct24_abs',3.0):.1f} | cooldown {a.get('cooldown_min',20)}м"
    )

# ---------------- COMMANDS & HANDLERS ----------------

@dp.message(Command("start"))
async def cmd_start(m: Message):
    ensure_user(m.from_user.id)
    header = await render_header_text()
    await m.answer(header + f"\n\nДобро пожаловать в <b>Innertrade Screener</b> {VERSION} (Bybit).",
                   reply_markup=bottom_menu_kb())


@dp.message(Command("status"))
async def cmd_status(m: Message):
    u = ensure_user(m.from_user.id)
    now = datetime.now(pytz.timezone(TZ)).strftime("%Y-%m-%d %H:%M:%S")
    wl = ", ".join(u["watchlist"]) if u["watchlist"] else "—"
    await m.answer(
        "<b>Status</b>\n"
        f"Time: {now} ({TZ})\n"
        f"Mode: {u['mode']} | Quiet: {u['quiet']}\n"
        "Source: Bybit (linear USDT)\n"
        f"Watchlist: {wl}\n"
        f"Webhook: ON\n"
        f"Version: {VERSION}"
    )


@dp.message(Command("diag"))
async def cmd_diag(m: Message):
    sym = "BTCUSDT"
    async with aiohttp.ClientSession() as s:
        k5 = await bybit_klines(s, sym, 5, 50)
        l5 = (k5 or {}).get("result", {}).get("list", [])
    await m.answer(f"Diag {sym}: 5m candles={len(l5)}")


@dp.message(F.text == "📊 Активность")
async def on_activity(m: Message):
    header = await render_header_text()
    body = await render_activity_text()
    await m.answer(header + "\n" + body, reply_markup=bottom_menu_kb())


@dp.message(F.text == "⚡ Волатильность")
async def on_vol(m: Message):
    header = await render_header_text()
    body = await render_volatility_text()
    await m.answer(header + "\n" + body, reply_markup=bottom_menu_kb())


@dp.message(F.text == "📈 Тренд")
async def on_trend(m: Message):
    header = await render_header_text()
    body = await render_trend_text()
    await m.answer(header + "\n" + body, reply_markup=bottom_menu_kb())


@dp.message(F.text == "📰 Новости")
async def on_news(m: Message):
    header = await render_header_text()
    await m.answer(header + "\n\n📰 <b>Макро (последний час)</b>\n• CPI (US) 3.1% vs 3.2% прогноз — риск-он")

Iurii Qechunts, [05.09.2025 23:54]
@dp.message(F.text == "🫧 Bubbles")
async def on_bubbles(m: Message):
    try:
        header, png = await render_bubbles_message()
        await m.answer_photo(photo=png, caption=header, reply_markup=bottom_menu_kb())
    except Exception as e:
        logging.warning(f"[BUBBLES ERR] {e}")
        await m.answer("Не удалось построить Bubbles (попробуйте позже).")


@dp.message(F.text == "🧮 Калькулятор")
async def on_calc(m: Message):
    try:
        data = await build_risk_excel_template()
        await m.answer_document(document=("risk_calc.xlsx", data), caption="Шаблон риск-менеджмента")
    except Exception as e:
        logging.warning(f"[EXCEL ERR] {e}")
        await m.answer("Не удалось сформировать Excel.")


@dp.message(F.text == "⭐ Watchlist")
async def on_watchlist_btn(m: Message):
    u = ensure_user(m.from_user.id)
    if not u["watchlist"]:
        await m.answer("Watchlist пуст. Добавь /add SYMBOL (например, /add SOLUSDT)")
    else:
        await m.answer("⭐ Watchlist:\n" + "\n".join(f"• {x}" for x in u["watchlist"]))


@dp.message(F.text == "⚙️ Настройки")
async def on_settings(m: Message):
    u = ensure_user(m.from_user.id)
    await m.answer(settings_text(u))


@dp.message(Command("add"))
async def cmd_add(m: Message):
    parts = m.text.split()
    if len(parts) < 2:
        await m.answer("Формат: /add SYMBOL (например /add SOLUSDT)")
        return
    sym = parts[1].upper()
    if not sym.endswith("USDT"):
        sym += "USDT"
    u = ensure_user(m.from_user.id)
    if sym not in SYMBOLS_BYBIT:
        SYMBOLS_BYBIT.append(sym)
    if sym not in u["watchlist"]:
        u["watchlist"].append(sym)
    await m.answer(f"Добавлено в Watchlist: {sym}")


@dp.message(Command("rm"))
async def cmd_rm(m: Message):
    parts = m.text.split()
    if len(parts) < 2:
        await m.answer("Формат: /rm SYMBOL (например /rm SOLUSDT)")
        return
    sym = parts[1].upper()
    u = ensure_user(m.from_user.id)
    if sym in u["watchlist"]:
        u["watchlist"].remove(sym)
        await m.answer(f"Удалено из Watchlist: {sym}")
    else:
        await m.answer(f"{sym} не найден в Watchlist")


@dp.message(Command("watchlist"))
async def cmd_watchlist(m: Message):
    u = ensure_user(m.from_user.id)
    if not u["watchlist"]:
        await m.answer("Watchlist пуст.")
    else:
        await m.answer("⭐ Watchlist:\n" + "\n".join(f"• {x}" for x in u["watchlist"]))


@dp.message(Command("passive"))
async def cmd_passive(m: Message):
    u = ensure_user(m.from_user.id)
    u["mode"] = "passive"
    await m.answer("Пассивный режим включен (автосигналы).")


@dp.message(Command("active"))
async def cmd_active(m: Message):
    u = ensure_user(m.from_user.id)
    u["mode"] = "active"
    await m.answer("Пассивный режим выключен.")


# ---------------- PASSIVE STREAM / ALERTS ----------------
async def fetch_watchlist_snapshot(symbols: list[str]) -> dict:
    out = {}
    async with aiohttp.ClientSession() as s:
        for sym in symbols:
            k5 = await bybit_klines(s, sym, 5, 50)
            tkr = await bybit_ticker(s, sym)
            l5 = (k5 or {}).get("result", {}).get("list", [])
            pct = 0.0
            try:
                lst = (tkr or {}).get("result", {}).get("list", [])
                if lst:
                    pct = float(lst[0].get("price24hPcnt", 0.0)) * 100.0
            except Exception:
                pass
            vol_mult = 0.0
            try:
                if len(l5) >= 25:
                    rows5 = [parse_bybit_row(r) for r in l5]
                    turns = [r[5] for r in rows5]
                    ma = moving_average(turns[:-1], 20)
                    if ma and ma > 0:
                        vol_mult = turns[-1] / ma
            except Exception:
                pass
            out[sym] = {"pct24": pct, "vol_mult": vol_mult}
    return out

Iurii Qechunts, [05.09.2025 23:54]
async def passive_loop():
    tz = pytz.timezone(TZ)
    while True:
        for uid, st in USERS.items():
            if st.get("mode") != "passive":
                continue
            if st.get("quiet"):
                now = datetime.now(tz).time()
                if 0 <= now.hour <= 7:
                    continue

            # 1) digest
            try:
                header = await render_header_text()
                body = await render_activity_text()
                await bot.send_message(uid, header + "\n" + body)
            except Exception as e:
                logging.warning(f"[PASSIVE DIGEST ERR] {e}")

            # 2) alerts by watchlist
            wl = st.get("watchlist", [])
            if wl:
                try:
                    snap = await fetch_watchlist_snapshot(wl)
                    a = st.get("alerts", {})
                    vm_thr = a.get("vol_mult", 1.5)
                    pct_thr = a.get("pct24_abs", 3.0)
                    cd_min = a.get("cooldown_min", 20)
                    now_ts = time()
                    for sym, vals in snap.items():
                        fire = False
                        reasons = []
                        if vals["vol_mult"] >= vm_thr:
                            fire = True; reasons.append(f"Vol x{vals['vol_mult']:.1f}≥{vm_thr:.1f}")
                        if abs(vals["pct24"]) >= pct_thr:
                            fire = True; reasons.append(f"|24h%| {abs(vals['pct24']):.1f}≥{pct_thr:.1f}")
                        if fire:
                            last_ts = st["last_alert"].get(sym, 0)
                            if now_ts - last_ts >= cd_min * 60:
                                st["last_alert"][sym] = now_ts
                                msg = f"🔔 <b>{sym}</b> — " + ", ".join(reasons)
                                with contextlib.suppress(Exception):
                                    await bot.send_message(uid, msg)
                except Exception as e:
                    logging.warning(f"[PASSIVE WL ERR] {e}")

        await asyncio.sleep(900)  # 15 minutes


# ---------------- AIOHTTP SERVER (WEBHOOK) ----------------
async def handle_health(request):
    return web.json_response({"ok": True, "service": "innertrade-screener", "version": VERSION})


async def handle_webhook(request):
    try:
        data = await request.json()
    except Exception:
        return web.Response(status=400)
    try:
        update = Update.model_validate(data)
        await dp.feed_update(bot, update)
    except Exception as e:
        logging.warning(f"[WEBHOOK ERR] {e}")
    return web.Response(status=200)


async def start_http_server():
    app = web.Application()
    app.router.add_get("/health", handle_health)
    app.router.add_post(f"/webhook/{TOKEN}", handle_webhook)
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.getenv("PORT", "10000"))  # Render injects PORT automatically
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    logging.info(f"HTTP server started on 0.0.0.0:{port}")


# ---------------- ENTRYPOINT ----------------
async def main():
    # start HTTP server
    await start_http_server()

    # reset webhook & set new
    try:
        await bot.delete_webhook(drop_pending_updates=True)
    except Exception:
        pass

    webhook_url = f"{BASE_URL}/webhook/{TOKEN}"
    allowed = dp.resolve_used_update_types()
    await bot.set_webhook(webhook_url, allowed_updates=allowed)
    logging.info(f"Webhook set to: {webhook_url}")

    # passive loop
    asyncio.create_task(passive_loop())

    # keep process alive
    while True:
        await asyncio.sleep(3600)


if name == "main":
    asyncio.run(main())